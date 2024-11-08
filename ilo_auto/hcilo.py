
"""
Purpose         : Generate HW healthcheck report from ILOM
Version         : 666.0
Author          : Shihab Istiak Saikat
Release Date    : 03-11-2024
"""

from datetime import datetime
import argparse
import pandas as pd
from pathlib import Path
import os
from hpilo import Ilo
from getpass import getpass
from glob import glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


formatted_date = datetime.now().strftime('%Y-%m-%d')
user_home = Path.home()
output_file_path = user_home / "Desktop" / f"ILO_HealthCheck_{formatted_date}.xlsx"
file_pattern = str(user_home / "Desktop" / "ILO_HealthCheck_*.xlsx")
matching_files = glob(file_pattern)

ilo_login_id = input("Enter ILO username: ")
ilo_password = getpass("Enter ILO password: ")


def main():
    invent_file = user_home / ".ilo_inventory"
    invent_dict, node_types = load_dict_from_file(invent_file)

    parser = argparse.ArgumentParser(prog='hcilo', description='Performs health check of ILO')
    parser.add_argument('--all', action='store_true', help='Perform health checks for all keys')
    parser.add_argument('key_name', nargs='*', type=str, choices=node_types, help="The specific key name to look for")
    args = parser.parse_args()

    terminal_width = os.get_terminal_size().columns
    completion_text = f"Successfully written."

    if args.all:
        remove_older_files()
        healthcheck(invent_dict)
        
        print(completion_text.center(terminal_width))
    elif args.key_name:
        remove_older_files()
        for key_name in args.key_name:
            healthcheck(invent_dict, key_name)
        print(completion_text.center(terminal_width))
    else:
        print("Please specify either --all or one or specify node types.")


def load_dict_from_file(invent_file):
    '''Load inventory from file into a dictionary.'''
    invent_dict = {}
    node_types = []
    current_key = None
    try:
        with open(invent_file, 'r') as file:
            for line in file:
                effective_line = line.strip()
                if effective_line:
                    if effective_line.endswith(':'):
                        current_key = effective_line[:-1]
                        invent_dict[current_key] = []
                        node_types.append(current_key)
                    else:
                        if current_key is not None:
                            ips = effective_line.split(',')
                            invent_dict[current_key].extend(ip.strip() for ip in ips if ip.strip())
                        else:
                            print("Warning: No current key set for line:", effective_line)

    except FileNotFoundError:
        print(f"Error: The file '{invent_file}' was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

    return invent_dict, node_types

# Check if older files exist
def remove_older_files():
    if matching_files:
        print("Older HealthCheck Report Found. Removing...")
        for file_path in matching_files:
            os.remove(file_path)
            print(f"Deleted old file: {file_path}")
    else:
        print("No older HealthCheck Report found. Continuing ...")

def health_check(ip, node_name, sheet):
    ilo = Ilo(ip, login=ilo_login_id, password=ilo_password)
    health = ilo.get_embedded_health()

    glance_data = health["health_at_a_glance"]
    glance_df = pd.DataFrame.from_dict(glance_data, orient='index')
    glance_df = glance_df.rename(columns={'status': 'Status', 'redundancy': 'Redundancy'}).rename_axis('Component').reset_index()
    glance_df.insert(0, 'Nodes', node_name)

    return glance_df

def write_to_excel(sheet_name, data):
    if os.path.exists(output_file_path):
        workbook = load_workbook(output_file_path)
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)
    else:
        workbook = Workbook()
        # Check if the default "Sheet" exists, as new workbooks have a default sheet
        if "Sheet" in workbook.sheetnames:
            worksheet = workbook["Sheet"]
            worksheet.title = sheet_name
        else:
            worksheet = workbook.create_sheet(sheet_name)

    # Define header styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Aqua color
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers only if the sheet is newly created or empty
    if worksheet.max_row == 1 and worksheet.max_column == 1:
        headers = list(data.columns)
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border

    # Find the starting row for appending data
    start_row = worksheet.max_row + 1
    
    # Write data to the worksheet
    for index, row in data.iterrows():
        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=start_row + index, column=col_num, value=value)
            cell.border = thin_border

    # Merge cells in the "Nodes" column for duplicate nodes
    node_col_idx = data.columns.get_loc("Nodes") + 1  # Column index for "Nodes" (1-based)
    start_row = worksheet.max_row - len(data) + 1  # Starting row for this batch of data
    merge_start = start_row
    last_node = worksheet.cell(row=start_row, column=node_col_idx).value

    for row in range(start_row + 1, start_row + len(data)):
        node_value = worksheet.cell(row=row, column=node_col_idx).value

        # Adjust the end row condition to prevent extra rows
        if node_value != last_node:
            if merge_start < row - 1:
                worksheet.merge_cells(start_row=merge_start, end_row=row - 1, start_column=node_col_idx, end_column=node_col_idx)
                worksheet.cell(row=merge_start, column=node_col_idx).alignment = center_alignment
            merge_start = row
        last_node = node_value

    # Final merge for the last group
    if merge_start < row:
        worksheet.merge_cells(start_row=merge_start, end_row=row, start_column=node_col_idx, end_column=node_col_idx)
        worksheet.cell(row=merge_start, column=node_col_idx).alignment = center_alignment

    workbook.save(output_file_path)


def healthcheck(invent_dict, key_name=None):
    if key_name:
        ips = invent_dict.get(key_name, [])
        for ip in ips:
            node_name = Ilo(ip, login=ilo_login_id, password=ilo_password).get_server_name()
            data = health_check(ip, node_name, key_name)
            write_to_excel(key_name, data)
            print(f"Data has written for {node_name} in '{output_file_path}'")
    else:  # No specific key, perform checks for all keys
        for key in invent_dict.keys():
            ips = invent_dict[key]
            for ip in ips:
                node_name = Ilo(ip, login=ilo_login_id, password=ilo_password).get_server_name()
                data = health_check(ip, node_name, key)
                write_to_excel(key, data)
                print(f"Data has been written for {node_name} in '{output_file_path}'")


if __name__ == '__main__':
    main()
